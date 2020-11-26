

#pd.set_option('max_columns', None)


from openpyxl import Workbook,load_workbook

wb = load_workbook("data.xlsx")
ws = wb.active


for i in range(1, 15828):
    try:
        """a = ws.cell(i, 11).value
        b = int(a.split(" ")[1].replace(".", ""))
        ws.cell(i, 11).value = b"""
        """if ws.cell(i, 9).value == 100001:
            ws.delete_rows(i)"""

    except:
        pass

wb.save("data.xlsx")


